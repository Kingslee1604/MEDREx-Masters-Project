"""
MEDREx — How Prediction Works (Excel Report)
Shows step by step how TF-IDF + LR predicts cancer type for a real report
Run: python generate_prediction_explainer.py
Output: prediction_explainer.xlsx
"""

import os, warnings
warnings.filterwarnings("ignore")

import joblib
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE    = r"c:\Users\kings\OneDrive\Desktop\CSUDH_Kingslee\CSUDH-Spring-2026\CSC-590-MastersProject\MastersProject"
REPORTS = os.path.join(BASE, "DataSet", "TCGA_Reports.csv", "TCGA_Reports.csv")
LABELS  = os.path.join(BASE, "DataSet", "tcga_patient_to_cancer_type.csv")
MODELS  = os.path.join(BASE, "backend", "models")
OUT     = os.path.join(BASE, "prediction_explainer.xlsx")

CANCER_NAMES = {
    "ACC":"Adrenocortical carcinoma","BLCA":"Bladder Urothelial Carcinoma",
    "BRCA":"Breast invasive carcinoma","CESC":"Cervical squamous cell carcinoma",
    "CHOL":"Cholangiocarcinoma","COAD":"Colon adenocarcinoma",
    "DLBC":"Diffuse Large B-cell Lymphoma","ESCA":"Esophageal carcinoma",
    "GBM":"Glioblastoma multiforme","HNSC":"Head and Neck squamous cell carcinoma",
    "KICH":"Kidney Chromophobe","KIRC":"Kidney renal clear cell carcinoma",
    "KIRP":"Kidney renal papillary cell carcinoma","LAML":"Acute Myeloid Leukemia",
    "LGG":"Brain Lower Grade Glioma","LIHC":"Liver hepatocellular carcinoma",
    "LUAD":"Lung adenocarcinoma","LUSC":"Lung squamous cell carcinoma",
    "MESO":"Mesothelioma","OV":"Ovarian serous cystadenocarcinoma",
    "PAAD":"Pancreatic adenocarcinoma","PCPG":"Pheochromocytoma and Paraganglioma",
    "PRAD":"Prostate adenocarcinoma","READ":"Rectum adenocarcinoma",
    "SARC":"Sarcoma","SKCM":"Skin Cutaneous Melanoma","STAD":"Stomach adenocarcinoma",
    "TGCT":"Testicular Germ Cell Tumors","THCA":"Thyroid carcinoma","THYM":"Thymoma",
    "UCEC":"Uterine Corpus Endometrial Carcinoma","UCS":"Uterine Carcinosarcoma",
    "UVM":"Uveal Melanoma",
}

# ── Styles ─────────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="1E293B", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def cw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write(ws, row, col, val, bg="FFFFFF", bold=False, fc="1E293B",
          size=11, align="left", italic=False, span_end=None, height=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.fill      = fill(bg)
    cell.font      = font(bold=bold, color=fc, size=size, italic=italic)
    cell.alignment = center() if align == "center" else left()
    cell.border    = border()
    if span_end:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=span_end)
    if height:
        ws.row_dimensions[row].height = height
    return cell

# ── Load data & model ──────────────────────────────────────────────────────────
print("Loading data and model...")
df_r = pd.read_csv(REPORTS)
df_l = pd.read_csv(LABELS)
df_r["patient_id"] = df_r["patient_filename"].apply(lambda x: x.split(".")[0])
df   = df_r.merge(df_l, on="patient_id", how="inner")

_, X_te, _, y_te = train_test_split(
    df["text"], df["cancer_type"],
    test_size=0.3, random_state=42, stratify=df["cancer_type"]
)

model         = joblib.load(os.path.join(MODELS, "tfidf_lr.pkl"))
vec           = joblib.load(os.path.join(MODELS, "tfidf_vec.pkl"))
feature_names = vec.get_feature_names_out()
classes       = model.classes_

# Pick one example per cancer type (correctly predicted)
X_te_v   = vec.transform(X_te)
y_pred   = model.predict(X_te_v)
probs_all = model.predict_proba(X_te_v)

# Pick 3 interesting examples: BRCA, GBM, THCA
examples = {}
for target in ["BRCA", "GBM", "THCA"]:
    idxs = [i for i,(t,p) in enumerate(zip(y_te.values, y_pred)) if t==target and p==target]
    if idxs:
        i = idxs[0]
        examples[target] = {
            "text":      X_te.values[i],
            "true":      y_te.values[i],
            "predicted": y_pred[i],
            "probs":     probs_all[i],
            "vec_row":   X_te_v[i],
        }

# ── Build Excel ────────────────────────────────────────────────────────────────
writer = pd.ExcelWriter(OUT, engine="openpyxl")
for sheet in ["How It Works", "Step-by-Step BRCA", "Step-by-Step GBM",
              "IDF Lookup Table", "Top Words per Cancer"]:
    pd.DataFrame().to_excel(writer, sheet_name=sheet, index=False)
writer.close()
wb = load_workbook(OUT)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — HOW IT WORKS (Overview)
# ══════════════════════════════════════════════════════════════════════════════
ws = wb["How It Works"]
ws.sheet_view.showGridLines = False

write(ws, 1, 1, "MEDREx — How TF-IDF + Logistic Regression Predicts Cancer Type",
      bg="1E3A5F", bold=True, fc="FFFFFF", size=16, align="center",
      span_end=6, height=36)
write(ws, 2, 1, "A step-by-step walkthrough of what happens inside the model when you paste a pathology report",
      bg="2D5A9E", fc="A8C8E8", size=11, italic=True, align="center",
      span_end=6, height=22)

# The two files explained
r = 4
write(ws, r, 1, "The Two Files That Make the Model", bg="E0E7F1", bold=True,
      fc="1E3A5F", size=13, span_end=6, align="center", height=24)

r += 1
for col, (val, bg_c) in enumerate([
    (".pkl File", "1E3A5F"),("What it is", "1E3A5F"),("Size", "1E3A5F"),
    ("What it learned", "1E3A5F"),("Used for", "1E3A5F"),("","1E3A5F")], 1):
    write(ws, r, col, val, bg=bg_c, bold=True, fc="FFFFFF", align="center", height=20)

r += 1
rows_files = [
    ("tfidf_vec.pkl", "TF-IDF Vectorizer",
     "~5 MB", "Vocabulary of 15,000 words + IDF weight for each word",
     "Converts any report text into a row of 15,000 numbers",
     "EFF6FF", "1E3A5F"),
    ("tfidf_lr.pkl", "Logistic Regression Model",
     "~15 MB", "525,000 weights (15,000 words × 35 cancer types)",
     "Scores each cancer type using those 15,000 numbers",
     "DCFCE7", "16A34A"),
]
for fname, ftype, fsize, learned, used, bg_c, fc_c in rows_files:
    for col, val in enumerate([fname, ftype, fsize, learned, used, ""], 1):
        write(ws, r, col, val, bg=bg_c, bold=(col==1), fc=fc_c if col==1 else "1E293B",
              height=24)
    r += 1

# The 4 steps
r += 1
write(ws, r, 1, "The 4 Steps That Happen Every Time You Predict",
      bg="E0E7F1", bold=True, fc="1E3A5F", size=13,
      span_end=6, align="center", height=24)
r += 1
for col, val in enumerate(["Step","Name","What Happens","Example Input","Example Output",""], 1):
    write(ws, r, col, val, bg="1E3A5F", bold=True, fc="FFFFFF", align="center", height=20)

steps = [
    ("1", "Clean & Tokenize",
     "Split report text into individual words and word pairs",
     '"Invasive ductal carcinoma, breast tissue..."',
     '["invasive", "ductal", "carcinoma", "breast", "invasive ductal", ...]',
     "EFF6FF"),
    ("2", "TF-IDF Lookup",
     "Look up each word in the vocabulary table. Multiply term frequency by IDF weight.",
     '"ductal" appears 3 times, IDF = 5.102',
     '"ductal" → TF-IDF score = 3 × 5.102 = 15.31',
     "FFF9C4"),
    ("3", "Score All 35 Cancer Types",
     "Multiply each word's TF-IDF score by its learned weight for each cancer type. Sum them all up.",
     'BRCA weight for "ductal" = +8.2',
     "BRCA total score = 187.4 | GBM total = -42.1 | LUAD = 23.6 | ...",
     "DCFCE7"),
    ("4", "Pick the Winner",
     "Convert all 35 scores to probabilities (0-100%). Pick the cancer type with highest probability.",
     "BRCA=92.3%, LUAD=3.1%, UCEC=1.8%, ...",
     "Prediction = BRCA (Breast invasive carcinoma) | Confidence = 92.3%",
     "DCFCE7"),
]
r += 1
for step, name, what, inp, out, bg_c in steps:
    for col, val in enumerate([step, name, what, inp, out, ""], 1):
        write(ws, r, col, val, bg=bg_c, bold=(col in [1,2]),
              fc="1E3A5F" if col in [1,2] else "1E293B", height=36)
    r += 1

# Key insight
r += 1
write(ws, r, 1,
      "Key Insight: The model never reads the text like a human. "
      "It converts text to numbers, multiplies by a table of 525,000 learned weights, "
      "and picks the highest score. All the 'intelligence' is stored in those weight numbers inside the .pkl file.",
      bg="FEF9C3", fc="92400E", bold=True, size=12,
      span_end=6, align="center", height=48)

for i in range(1, 7):
    cw(ws, i, [6, 22, 35, 35, 45, 4][i-1])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Step-by-Step BRCA Example
# ══════════════════════════════════════════════════════════════════════════════
for sheet_name, cancer_code in [("Step-by-Step BRCA", "BRCA"), ("Step-by-Step GBM", "GBM")]:
    ws2 = wb[sheet_name]
    ws2.sheet_view.showGridLines = False

    ex   = examples.get(cancer_code, examples[list(examples.keys())[0]])
    code = ex["true"]
    name = CANCER_NAMES.get(code, code)
    text = ex["text"][:500]
    probs = ex["probs"]
    vec_row = ex["vec_row"]

    write(ws2, 1, 1, f"Step-by-Step Prediction — Real {code} Report",
          bg="1E3A5F", bold=True, fc="FFFFFF", size=16,
          span_end=5, align="center", height=36)
    write(ws2, 2, 1, f"Actual cancer type: {code} — {name}   |   Model prediction: CORRECT",
          bg="16A34A", fc="FFFFFF", size=12, bold=True,
          span_end=5, align="center", height=24)

    # The actual report text
    r = 4
    write(ws2, r, 1, "ACTUAL REPORT TEXT (first 500 characters)",
          bg="E0E7F1", bold=True, fc="1E3A5F", size=12,
          span_end=5, align="center", height=22)
    r += 1
    ws2.merge_cells(f"A{r}:E{r}")
    cell = ws2.cell(row=r, column=1, value=text + "...")
    cell.fill = fill("F8FAFC")
    cell.font = Font(name="Calibri", size=10, italic=True, color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = border()
    ws2.row_dimensions[r].height = 80
    r += 2

    # Step 1 — top words found in this report
    write(ws2, r, 1, "STEP 1 — Top 20 Words Found in This Report (with TF-IDF scores)",
          bg="E0E7F1", bold=True, fc="1E3A5F", size=12,
          span_end=5, align="center", height=22)
    r += 1
    for col, val in enumerate(["Rank","Word / Phrase","TF-IDF Score",
                                "How rare is this word?",""], 1):
        write(ws2, r, col, val, bg="1E3A5F", bold=True, fc="FFFFFF",
              align="center", height=20)
    r += 1

    # Get top words for this specific report
    row_arr = vec_row.toarray()[0]
    top_idx = np.argsort(row_arr)[::-1][:20]
    idf_vals = vec.idf_

    for rank, idx in enumerate(top_idx, 1):
        score  = row_arr[idx]
        word   = feature_names[idx]
        idf    = idf_vals[idx]
        if idf > 6:   rarity = "Very rare — highly specific"
        elif idf > 4: rarity = "Rare — somewhat specific"
        elif idf > 2: rarity = "Common — moderate value"
        else:         rarity = "Very common — low value"

        bg_c = "DCFCE7" if idf > 5 else ("FFF9C4" if idf > 3 else "FEE2E2")
        for col, val in enumerate([rank, word, round(score,4), rarity, ""], 1):
            write(ws2, r, col, val, bg=bg_c,
                  bold=(col==2), fc="1E293B", align="center" if col!=2 else "left",
                  height=18)
        r += 1

    r += 1

    # Step 2 — how this word scores for each cancer
    write(ws2, r, 1, "STEP 2 — How the Top Word Scores Against All 35 Cancer Types",
          bg="E0E7F1", bold=True, fc="1E3A5F", size=12,
          span_end=5, align="center", height=22)
    r += 1
    top_word    = feature_names[top_idx[0]]
    top_word_i  = top_idx[0]
    write(ws2, r, 1, f'Example word: "{top_word}" (highest TF-IDF score in this report)',
          bg="FFF9C4", fc="92400E", bold=True, size=11,
          span_end=5, align="center", height=22)
    r += 1
    for col, val in enumerate(["Cancer Code","Cancer Name",
                                f'LR Weight for "{top_word}"',
                                "Interpretation",""], 1):
        write(ws2, r, col, val, bg="1E3A5F", bold=True, fc="FFFFFF",
              align="center", height=20)
    r += 1

    word_weights = [(cls, model.coef_[i][top_word_i])
                    for i, cls in enumerate(classes)]
    word_weights.sort(key=lambda x: x[1], reverse=True)

    for cls, w in word_weights:
        if w > 1:    bg_c, interp = "DCFCE7", "Strong POSITIVE signal for this cancer"
        elif w > 0:  bg_c, interp = "F0FDF4", "Weak positive signal"
        elif w > -1: bg_c, interp = "FEF9C3", "Weak negative signal"
        else:        bg_c, interp = "FEE2E2", "Strong NEGATIVE — word argues AGAINST this cancer"
        highlight = cls == code
        for col, val in enumerate([cls, CANCER_NAMES.get(cls,cls),
                                    round(w,4), interp, ""], 1):
            cell = write(ws2, r, col, val, bg="DCFCE7" if highlight else bg_c,
                         bold=highlight, fc="16A34A" if highlight else "1E293B",
                         align="center" if col in [1,3] else "left", height=18)
        r += 1

    r += 1

    # Step 3 — Final scores for all 35 types
    write(ws2, r, 1, "STEP 3 — Final Probability for All 35 Cancer Types (after summing ALL words)",
          bg="E0E7F1", bold=True, fc="1E3A5F", size=12,
          span_end=5, align="center", height=22)
    r += 1
    for col, val in enumerate(["Rank","Cancer Code","Cancer Name",
                                "Probability","Result"], 1):
        write(ws2, r, col, val, bg="1E3A5F", bold=True, fc="FFFFFF",
              align="center", height=20)
    r += 1

    sorted_probs = sorted(zip(classes, probs), key=lambda x: x[1], reverse=True)
    for rank, (cls, prob) in enumerate(sorted_probs, 1):
        is_winner = cls == code
        bg_c = "DCFCE7" if is_winner else ("FFF9C4" if rank <= 3 else "FFFFFF")
        result = "WINNER - PREDICTION" if is_winner else ("Top 3" if rank <= 3 else "")
        for col, val in enumerate([rank, cls, CANCER_NAMES.get(cls,cls),
                                    f"{prob*100:.2f}%", result], 1):
            write(ws2, r, col, val, bg=bg_c,
                  bold=is_winner, fc="16A34A" if is_winner else "1E293B",
                  align="center" if col in [1,2,4] else "left", height=18)
        r += 1

    for i, w in enumerate([8, 10, 32, 14, 20], 1):
        cw(ws2, i, w)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — IDF Lookup Table (sample of vocabulary)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb["IDF Lookup Table"]
ws3.sheet_view.showGridLines = False

write(ws3, 1, 1, "IDF Lookup Table — Sample from tfidf_vec.pkl",
      bg="1E3A5F", bold=True, fc="FFFFFF", size=15,
      span_end=5, align="center", height=34)
write(ws3, 2, 1,
      "This table shows what the model LEARNED about each word during training. "
      "High IDF = rare word = important. Low IDF = common word = ignored.",
      bg="2D5A9E", fc="A8C8E8", size=11, italic=True,
      span_end=5, align="center", height=24)
write(ws3, 3, 1, f"Total vocabulary size: {len(feature_names):,} words and word pairs",
      bg="E0E7F1", bold=True, fc="1E3A5F", size=12,
      span_end=5, align="center", height=22)

r = 5
for col, val in enumerate(["Rank","Word / Phrase","IDF Weight",
                            "Rarity","What This Means"], 1):
    write(ws3, r, col, val, bg="1E3A5F", bold=True, fc="FFFFFF",
          align="center", height=20)

# Sort by IDF descending — most specific words first
idf_vals   = vec.idf_
top_by_idf = np.argsort(idf_vals)[::-1]

# Show: top 30 most specific + bottom 20 most common
selected = list(top_by_idf[:30]) + list(np.argsort(idf_vals)[:20])
labels_used = []

r += 1
section_done = False
for rank, idx in enumerate(selected, 1):
    word = feature_names[idx]
    idf  = idf_vals[idx]

    if rank == 31 and not section_done:
        write(ws3, r, 1, "--- Most Common Words (Low IDF = low importance) ---",
              bg="FEE2E2", bold=True, fc="DC2626",
              span_end=5, align="center", height=20)
        r += 1
        section_done = True

    if idf > 7:
        bg_c  = "DCFCE7"; rarity = "Extremely rare — very cancer-specific"
        fc_c  = "16A34A"
    elif idf > 5:
        bg_c  = "F0FDF4"; rarity = "Rare — highly informative"
        fc_c  = "1E293B"
    elif idf > 3:
        bg_c  = "FFF9C4"; rarity = "Moderate — somewhat useful"
        fc_c  = "1E293B"
    else:
        bg_c  = "FEE2E2"; rarity = "Very common — mostly ignored"
        fc_c  = "DC2626"

    meaning = (
        "Appears in very few cancer types — strong signal" if idf > 6 else
        "Appears in a handful of cancer types" if idf > 4 else
        "Appears in many cancer types — weak signal" if idf > 2 else
        "Appears in almost every report — near useless for prediction"
    )

    for col, val in enumerate([rank, word, round(idf,4), rarity, meaning], 1):
        write(ws3, r, col, val, bg=bg_c,
              bold=(col==2), fc=fc_c if col in [2,3,4] else "1E293B",
              align="center" if col in [1,3] else "left", height=18)
    r += 1

for i, w in enumerate([6, 28, 12, 30, 42], 1):
    cw(ws3, i, w)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Top Words per Cancer Type
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb["Top Words per Cancer"]
ws4.sheet_view.showGridLines = False

write(ws4, 1, 1, "Top Predictive Words per Cancer Type — From tfidf_lr.pkl",
      bg="1E3A5F", bold=True, fc="FFFFFF", size=15,
      span_end=6, align="center", height=34)
write(ws4, 2, 1,
      "These are the words that MOST strongly push the model toward each cancer type. "
      "High positive weight = strong evidence FOR that cancer. "
      "These were learned automatically from 6,666 training reports.",
      bg="2D5A9E", fc="A8C8E8", size=11, italic=True,
      span_end=6, align="center", height=28)

r = 4
for col, val in enumerate(["Code","Cancer Name","#1 Word","#2 Word",
                            "#3 Word","#4 Word","#5 Word"], 1):
    write(ws4, r, col, val, bg="1E3A5F", bold=True, fc="FFFFFF",
          align="center", height=22)

row_colors = ["EFF6FF","F0FDF4","FFF9C4","FEF3C7","F5F3FF",
              "EFF6FF","F0FDF4","FFF9C4","FEF3C7","F5F3FF"]

r += 1
for ci, cls in enumerate(sorted(classes)):
    coef  = model.coef_[list(classes).index(cls)]
    top5i = np.argsort(coef)[::-1][:5]
    top5w = [feature_names[j] for j in top5i]
    bg_c  = row_colors[ci % len(row_colors)]
    vals  = [cls, CANCER_NAMES.get(cls, cls)] + top5w
    for col, val in enumerate(vals, 1):
        write(ws4, r, col, val, bg=bg_c,
              bold=(col <= 2), fc="1E3A5F" if col <= 2 else "1E293B",
              align="center" if col == 1 else "left", height=20)
    r += 1

for i, w in enumerate([8, 35, 22, 22, 22, 22, 22], 1):
    cw(ws4, i, w)


# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"\nSaved: {OUT}")
print("5 sheets:")
print("  Sheet 1 - How It Works         (overview of all 4 steps)")
print("  Sheet 2 - Step-by-Step BRCA    (real report walked through completely)")
print("  Sheet 3 - Step-by-Step GBM     (real report walked through completely)")
print("  Sheet 4 - IDF Lookup Table     (sample of 50 words from the vocabulary)")
print("  Sheet 5 - Top Words per Cancer (what word = what cancer)")
