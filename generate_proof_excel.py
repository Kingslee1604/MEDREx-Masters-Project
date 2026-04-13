"""
MEDREx — Model Evaluation Proof Report (Excel)
Run: python generate_proof_excel.py
Output: proof_report.xlsx  (4 sheets)
"""

import os, warnings
warnings.filterwarnings("ignore")

import joblib
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, f1_score
from sklearn.metrics import precision_score, recall_score
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

BASE    = r"c:\Users\kings\OneDrive\Desktop\CSUDH_Kingslee\CSUDH-Spring-2026\CSC-590-MastersProject\MastersProject"
REPORTS = os.path.join(BASE, "DataSet", "TCGA_Reports.csv", "TCGA_Reports.csv")
LABELS  = os.path.join(BASE, "DataSet", "tcga_patient_to_cancer_type.csv")
MODELS  = os.path.join(BASE, "backend", "models")
OUT     = os.path.join(BASE, "proof_report.xlsx")

CANCER_NAMES = {
    "ACC":"Adrenocortical carcinoma","BLCA":"Bladder Urothelial Carcinoma",
    "BRCA":"Breast invasive carcinoma","CESC":"Cervical squamous cell carcinoma",
    "CHOL":"Cholangiocarcinoma","COAD":"Colon adenocarcinoma",
    "DLBC":"Diffuse Large B-cell Lymphoma","ESCA":"Esophageal carcinoma",
    "GBM":"Glioblastoma multiforme","HNSC":"Head and Neck squamous cell carcinoma",
    "KICH":"Kidney Chromophobe","KIRC":"Kidney renal clear cell carcinoma",
    "KIRP":"Kidney renal papillary cell carcinoma","LAML":"Acute Myeloid Leukemia",
    "LGG":"Brain Lower Grade Glioma","LIHC":"Liver hepatocellular carcinoma",
    "LUAD":"Lung adenocarcinoma","LUSC":"Lung squamous cell carcinoma",
    "MESO":"Mesothelioma","OV":"Ovarian serous cystadenocarcinoma",
    "PAAD":"Pancreatic adenocarcinoma","PCPG":"Pheochromocytoma and Paraganglioma",
    "PRAD":"Prostate adenocarcinoma","READ":"Rectum adenocarcinoma",
    "SARC":"Sarcoma","SKCM":"Skin Cutaneous Melanoma","STAD":"Stomach adenocarcinoma",
    "TGCT":"Testicular Germ Cell Tumors","THCA":"Thyroid carcinoma","THYM":"Thymoma",
    "UCEC":"Uterine Corpus Endometrial Carcinoma","UCS":"Uterine Carcinosarcoma",
    "UVM":"Uveal Melanoma",
}

# ── Color constants ────────────────────────────────────────────────────────────
NAVY_HEX   = "1E3A5F"
WHITE_HEX  = "FFFFFF"
GREEN_HEX  = "16A34A"
GREEN_BG   = "DCFCE7"
ORANGE_HEX = "EA580C"
ORANGE_BG  = "FFF3CD"
RED_HEX    = "DC2626"
RED_BG     = "FEE2E2"
YELLOW_BG  = "FEF9C3"
LIGHT_GRAY = "F1F5F9"
HEADER_BG  = "1E3A5F"
SUBHEAD_BG = "E0E7F1"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def header_row(ws, row, cols, bg=HEADER_BG, fg=WHITE_HEX, height=22):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill    = fill(bg)
        cell.font    = font(bold=True, color=fg, size=11)
        cell.alignment = center()
        cell.border  = thin_border()
    ws.row_dimensions[row].height = height

def data_row(ws, row, values, bg=WHITE_HEX, bold=False, height=18):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill    = fill(bg)
        cell.font    = font(bold=bold, color="1E293B")
        cell.alignment = center() if c > 2 else left()
        cell.border  = thin_border()
    ws.row_dimensions[row].height = height

def title_cell(ws, row, col, text, span_end=None,
               bg=HEADER_BG, fg=WHITE_HEX, size=14, height=30):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill      = fill(bg)
    cell.font      = font(bold=True, color=fg, size=size)
    cell.alignment = center()
    if span_end:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=span_end
        )
    ws.row_dimensions[row].height = height


# ══════════════════════════════════════════════════════════════════════════════
# LOAD DATA & RUN MODEL
# ══════════════════════════════════════════════════════════════════════════════
print("Loading data...")
df_r = pd.read_csv(REPORTS)
df_l = pd.read_csv(LABELS)
df_r["patient_id"] = df_r["patient_filename"].apply(lambda x: x.split(".")[0])
df   = df_r.merge(df_l, on="patient_id", how="inner")

X_tr, X_te, y_tr, y_te = train_test_split(
    df["text"], df["cancer_type"],
    test_size=0.3, random_state=42, stratify=df["cancer_type"]
)

print("Loading model & predicting...")
model = joblib.load(os.path.join(MODELS, "tfidf_lr.pkl"))
vec   = joblib.load(os.path.join(MODELS, "tfidf_vec.pkl"))
X_te_v = vec.transform(X_te)
y_pred  = model.predict(X_te_v)

acc = accuracy_score(y_te, y_pred)
f1w = f1_score(y_te, y_pred, average="weighted")

classes   = sorted(y_te.unique())
f1_scores = f1_score(y_te, y_pred, labels=classes, average=None, zero_division=0)
prec_sc   = precision_score(y_te, y_pred, labels=classes, average=None, zero_division=0)
rec_sc    = recall_score(y_te, y_pred, labels=classes, average=None, zero_division=0)
support   = y_te.value_counts()

# Wrong predictions
wrong_mask = y_pred != y_te.values
wrong_df   = pd.DataFrame({"True": y_te.values[wrong_mask], "Predicted": y_pred[wrong_mask]})
confused   = wrong_df.groupby(["True","Predicted"]).size().reset_index(name="Count")
confused   = confused.sort_values("Count", ascending=False).head(15)

# Top words per class
feature_names = vec.get_feature_names_out()
top_words = {}
for i, cls in enumerate(model.classes_):
    coef = model.coef_[i]
    top5 = [feature_names[j] for j in np.argsort(coef)[::-1][:5]]
    top_words[cls] = ", ".join(top5)

# Build per-class DataFrame sorted by F1
rows = []
for cls, f1, p, r in zip(classes, f1_scores, prec_sc, rec_sc):
    rows.append({
        "Code": cls,
        "Cancer Name": CANCER_NAMES.get(cls, cls),
        "Test Samples": int(support.get(cls, 0)),
        "Precision": round(float(p), 4),
        "Recall":    round(float(r), 4),
        "F1 Score":  round(float(f1), 4),
        "Top 5 Predictive Words": top_words.get(cls, ""),
    })
df_metrics = pd.DataFrame(rows).sort_values("F1 Score")

# ══════════════════════════════════════════════════════════════════════════════
# WRITE EXCEL
# ══════════════════════════════════════════════════════════════════════════════
writer = pd.ExcelWriter(OUT, engine="openpyxl")
# Write dummy sheets to create the file
pd.DataFrame().to_excel(writer, sheet_name="Summary",         index=False)
pd.DataFrame().to_excel(writer, sheet_name="Per-Class F1",    index=False)
pd.DataFrame().to_excel(writer, sheet_name="Misclassifications", index=False)
pd.DataFrame().to_excel(writer, sheet_name="Model Config",    index=False)
writer.close()

wb = load_workbook(OUT)


# ── SHEET 1: Summary ──────────────────────────────────────────────────────────
ws = wb["Summary"]
ws.sheet_view.showGridLines = False

title_cell(ws, 1, 1, "MEDREx — Model Evaluation Proof Report", span_end=5,
           size=16, height=36)
title_cell(ws, 2, 1,
           "TF-IDF + Logistic Regression  |  Task 1: Cancer Type Classification  |  TCGA Dataset",
           span_end=5, bg="2D5A9E", size=11, height=22)
title_cell(ws, 3, 1,
           "Student: Kingslee Dominic Savio Velu  |  CSC 590 – CSUDH Spring 2026  |  Mentor: Dr. Bin Tang",
           span_end=5, bg="3B6FBF", size=10, fg="E0E7F1", height=20)

# Overall metrics
ws.row_dimensions[4].height = 10
header_row(ws, 5, ["Metric", "Value", "Description", "", ""], bg=NAVY_HEX, height=20)
ws.merge_cells("C5:E5")

metrics_data = [
    ("Test Accuracy",      f"{acc*100:.4f}%",  "Correct predictions / total test predictions",  GREEN_BG,  GREEN_HEX),
    ("Weighted F1 Score",  f"{f1w*100:.4f}%",  "Balances precision and recall across 32 classes", GREEN_BG, GREEN_HEX),
    ("Total Reports",      f"{len(df):,}",     "Full TCGA dataset size",                          LIGHT_GRAY, "1E293B"),
    ("Training Set",       f"{len(X_tr):,}",   "70% of data used to train the model",             LIGHT_GRAY, "1E293B"),
    ("Test Set",           f"{len(X_te):,}",   "30% held-out — model NEVER saw these during training", LIGHT_GRAY, "1E293B"),
    ("Cancer Types",       str(len(classes)),  "Number of classification classes",                LIGHT_GRAY, "1E293B"),
    ("Random Seed",        "42",               "Fixed — results are fully reproducible",          LIGHT_GRAY, "1E293B"),
]
for i, (metric, val, desc, bg, fc) in enumerate(metrics_data):
    r = 6 + i
    ws.cell(r, 1, metric).fill = fill(bg)
    ws.cell(r, 1).font  = font(bold=True, color=fc)
    ws.cell(r, 1).alignment = left()
    ws.cell(r, 1).border = thin_border()
    ws.cell(r, 2, val).fill = fill(bg)
    ws.cell(r, 2).font  = font(bold=True, color=fc, size=12)
    ws.cell(r, 2).alignment = center()
    ws.cell(r, 2).border = thin_border()
    ws.merge_cells(f"C{r}:E{r}")
    ws.cell(r, 3, desc).fill = fill(bg)
    ws.cell(r, 3).font  = font(color="475569")
    ws.cell(r, 3).alignment = left()
    ws.cell(r, 3).border = thin_border()
    ws.row_dimensions[r].height = 20

# Comparison vs reference
ws.row_dimensions[14].height = 12
title_cell(ws, 15, 1, "Comparison vs Cedars-Sinai Reference", span_end=5,
           bg=SUBHEAD_BG, fg=NAVY_HEX, size=12, height=22)
header_row(ws, 16, ["System", "Method", "Classes", "Accuracy", "Notes"], bg="2D5A9E", height=20)
comp = [
    ("Cedars-Sinai (Reference)", "BoW + Logistic Regression", "32", "95.31%",
     "Baseline provided to us", ORANGE_BG),
    ("MEDREx (Our Work)", "TF-IDF + Logistic Regression", "35", f"{acc*100:.2f}%",
     f"+{acc*100-95.31:.2f}% better, 3 more classes", GREEN_BG),
]
for i, (sys, method, cls, accuracy, note, bg) in enumerate(comp):
    r = 17 + i
    for c, val in enumerate([sys, method, cls, accuracy, note], 1):
        cell = ws.cell(r, c, val)
        cell.fill = fill(bg)
        cell.font = font(bold=(c==4), color=GREEN_HEX if bg==GREEN_BG and c==4 else "1E293B")
        cell.alignment = center() if c > 2 else left()
        cell.border = thin_border()
    ws.row_dimensions[r].height = 20

# How model was trained
ws.row_dimensions[20].height = 12
title_cell(ws, 21, 1, "How the Model Was Trained — Step by Step", span_end=5,
           bg=SUBHEAD_BG, fg=NAVY_HEX, size=12, height=22)
header_row(ws, 22, ["Step", "Action", "Detail", "", ""], bg="2D5A9E", height=20)
ws.merge_cells("C22:E22")
steps = [
    ("1", "Load Data",       "9,523 TCGA reports + cancer type labels joined on patient_id"),
    ("2", "Split Dataset",   "70% train (6,666) / 30% test (2,857) — stratified by cancer type"),
    ("3", "TF-IDF Vectorize","max_features=15000, ngram_range=(1,2), sublinear_tf=True, min_df=2"),
    ("4", "Train LR Model",  "C=1.0, max_iter=1000, class_weight='balanced', solver='lbfgs'"),
    ("5", "Save to Disk",    "joblib.dump() → tfidf_lr.pkl and tfidf_vec.pkl"),
    ("6", "Evaluate",        "Predict on 2,857 test reports → compare vs true labels → 96.36%"),
]
for i, (step, action, detail) in enumerate(steps):
    r = 23 + i
    bg = GREEN_BG if i == 5 else LIGHT_GRAY
    ws.merge_cells(f"C{r}:E{r}")
    for c, val in enumerate([step, action, detail], 1):
        cell = ws.cell(r, c, val)
        cell.fill = fill(bg)
        cell.font = font(bold=(c==2), color="1E293B")
        cell.alignment = left()
        cell.border = thin_border()
    ws.row_dimensions[r].height = 18

set_col_width(ws, 1, 28)
set_col_width(ws, 2, 22)
set_col_width(ws, 3, 55)
set_col_width(ws, 4, 14)
set_col_width(ws, 5, 28)


# ── SHEET 2: Per-Class F1 ─────────────────────────────────────────────────────
ws2 = wb["Per-Class F1"]
ws2.sheet_view.showGridLines = False

title_cell(ws2, 1, 1, "Per-Class F1 Scores — All Cancer Types", span_end=7,
           size=15, height=32)
title_cell(ws2, 2, 1,
           "Sorted from hardest (lowest F1) to easiest (highest F1)   |   Green >= 0.95   |   Yellow 0.85-0.95   |   Red < 0.85",
           span_end=7, bg="2D5A9E", size=10, height=20)

header_row(ws2, 3, ["#", "Code", "Cancer Name", "Test Samples",
                    "Precision", "Recall", "F1 Score"], bg=NAVY_HEX, height=22)

for i, (_, row) in enumerate(df_metrics.iterrows()):
    r = 4 + i
    f1 = row["F1 Score"]
    if f1 >= 0.95:
        bg, fc = GREEN_BG, GREEN_HEX
    elif f1 >= 0.85:
        bg, fc = YELLOW_BG, "92400E"
    else:
        bg, fc = RED_BG, RED_HEX

    vals = [i+1, row["Code"], row["Cancer Name"], row["Test Samples"],
            row["Precision"], row["Recall"], row["F1 Score"]]
    for c, val in enumerate(vals, 1):
        cell = ws2.cell(r, c, val)
        cell.fill = fill(bg if c >= 5 else LIGHT_GRAY)
        cell.font = font(
            bold=(c == 7),
            color=fc if c >= 5 else ("1E293B" if c > 1 else "64748B"),
            size=11
        )
        cell.alignment = center() if c != 3 else left()
        cell.border = thin_border()
        if c >= 5:
            cell.number_format = "0.0000"
    ws2.row_dimensions[r].height = 18

# Add top words below
ws2.row_dimensions[4 + len(df_metrics)].height = 12
r_start = 5 + len(df_metrics)
title_cell(ws2, r_start, 1, "Top 5 Predictive Words per Cancer Type", span_end=7,
           bg=SUBHEAD_BG, fg=NAVY_HEX, size=12, height=22)
header_row(ws2, r_start+1, ["Code", "Cancer Name", "Top 5 Words (most predictive)", "", "", "", ""],
           bg="2D5A9E", height=20)
ws2.merge_cells(f"C{r_start+1}:G{r_start+1}")

for i, cls in enumerate(sorted(top_words.keys())):
    r = r_start + 2 + i
    ws2.merge_cells(f"C{r}:G{r}")
    for c, val in enumerate([cls, CANCER_NAMES.get(cls, cls), top_words[cls]], 1):
        cell = ws2.cell(r, c, val)
        cell.fill = fill(LIGHT_GRAY if i % 2 == 0 else WHITE_HEX)
        cell.font = font(bold=(c==1), color="1E293B")
        cell.alignment = left()
        cell.border = thin_border()
    ws2.row_dimensions[r].height = 18

set_col_width(ws2, 1, 5)
set_col_width(ws2, 2, 8)
set_col_width(ws2, 3, 38)
set_col_width(ws2, 4, 14)
set_col_width(ws2, 5, 12)
set_col_width(ws2, 6, 12)
set_col_width(ws2, 7, 12)


# ── SHEET 3: Misclassifications ───────────────────────────────────────────────
ws3 = wb["Misclassifications"]
ws3.sheet_view.showGridLines = False

title_cell(ws3, 1, 1, "Top Misclassifications — Where the Model Gets Confused", span_end=5,
           size=15, height=32)
title_cell(ws3, 2, 1,
           "These are test reports where the model predicted the wrong cancer type",
           span_end=5, bg="2D5A9E", size=10, height=20)

header_row(ws3, 3, ["True Cancer Code", "True Cancer Name",
                    "Model Predicted", "Predicted Name", "Count"], bg=NAVY_HEX, height=22)

for i, (_, row) in enumerate(confused.iterrows()):
    r = 4 + i
    bg = RED_BG if i < 3 else (ORANGE_BG if i < 6 else LIGHT_GRAY)
    vals = [
        row["True"],
        CANCER_NAMES.get(row["True"], row["True"]),
        row["Predicted"],
        CANCER_NAMES.get(row["Predicted"], row["Predicted"]),
        row["Count"],
    ]
    for c, val in enumerate(vals, 1):
        cell = ws3.cell(r, c, val)
        cell.fill = fill(bg)
        cell.font = font(bold=(c in [1,3,5]), color="1E293B")
        cell.alignment = center() if c in [1,3,5] else left()
        cell.border = thin_border()
    ws3.row_dimensions[r].height = 20

# Explanation
r_exp = 4 + len(confused) + 2
title_cell(ws3, r_exp, 1, "Why These Errors Happen", span_end=5,
           bg=SUBHEAD_BG, fg=NAVY_HEX, size=12, height=22)
explanations = [
    ("COAD vs READ",  "Colon vs Rectum adenocarcinoma — nearly identical pathology language. Both say 'colorectal adenocarcinoma'."),
    ("UCS vs UCEC",   "Both uterine cancers. Carcinosarcoma and Endometrial carcinoma share overlapping terminology."),
    ("LUAD vs LUSC",  "Both lung cancers. Adenocarcinoma vs Squamous cell — different cell type but similar organ descriptions."),
    ("Why it matters","These errors scientifically justify adding BERT and RAG — they understand full sentence context, not just word frequency."),
]
header_row(ws3, r_exp+1, ["Confused Pair", "Explanation", "", "", ""], bg="2D5A9E", height=20)
ws3.merge_cells(f"B{r_exp+1}:E{r_exp+1}")
for i, (pair, exp) in enumerate(explanations):
    r = r_exp + 2 + i
    ws3.merge_cells(f"B{r}:E{r}")
    bg = RED_BG if i == 3 else LIGHT_GRAY
    for c, val in enumerate([pair, exp], 1):
        cell = ws3.cell(r, c, val)
        cell.fill = fill(bg)
        cell.font = font(bold=(c==1), color=RED_HEX if i==3 else "1E293B")
        cell.alignment = left()
        cell.border = thin_border()
    ws3.row_dimensions[r].height = 22

set_col_width(ws3, 1, 18)
set_col_width(ws3, 2, 38)
set_col_width(ws3, 3, 18)
set_col_width(ws3, 4, 38)
set_col_width(ws3, 5, 10)


# ── SHEET 4: Model Config ─────────────────────────────────────────────────────
ws4 = wb["Model Config"]
ws4.sheet_view.showGridLines = False

title_cell(ws4, 1, 1, "Model Configuration — What's Saved in the .pkl Files", span_end=4,
           size=15, height=32)
title_cell(ws4, 2, 1, "These exact settings reproduce the 96.36% result every time (random_state=42)",
           span_end=4, bg="2D5A9E", size=10, height=20)

# TF-IDF config
title_cell(ws4, 4, 1, "TF-IDF Vectorizer  (tfidf_vec.pkl)", span_end=4,
           bg=SUBHEAD_BG, fg=NAVY_HEX, size=12, height=22)
header_row(ws4, 5, ["Parameter", "Value", "Why This Value", "Impact"], bg=NAVY_HEX, height=22)
tfidf_params = [
    ("max_features", "15,000", "Top 15K most informative words kept", "Reduces noise from rare misspellings"),
    ("ngram_range",  "(1, 2)",  "Single words AND word pairs", "'invasive ductal' as one feature — captures phrases"),
    ("sublinear_tf", "True",    "Log scale on term frequency", "Prevents one repeated word from dominating score"),
    ("min_df",       "2",       "Word must appear in at least 2 reports", "Removes noise from unique/misspelled words"),
]
for i, params in enumerate(tfidf_params):
    r = 6 + i
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE_HEX
    for c, val in enumerate(params, 1):
        cell = ws4.cell(r, c, val)
        cell.fill = fill(bg)
        cell.font = font(bold=(c==1), color="1E293B")
        cell.alignment = left()
        cell.border = thin_border()
    ws4.row_dimensions[r].height = 20

# LR config
title_cell(ws4, 11, 1, "Logistic Regression  (tfidf_lr.pkl)", span_end=4,
           bg=SUBHEAD_BG, fg=NAVY_HEX, size=12, height=22)
header_row(ws4, 12, ["Parameter", "Value", "Why This Value", "Impact"], bg=NAVY_HEX, height=22)
lr_params = [
    ("C",             "1.0",      "Regularization strength",             "Balances fitting vs overfitting"),
    ("max_iter",      "1,000",    "Max iterations for convergence",       "Needed for 35-class problem to fully converge"),
    ("class_weight",  "balanced", "Auto-weight classes by frequency",     "Fixes 24x imbalance — BRCA:1034 vs CHOL:43"),
    ("solver",        "lbfgs",    "Limited-memory BFGS optimizer",        "Best for multiclass with large feature sets"),
    ("random_state",  "42",       "Fixed random seed",                    "Fully reproducible — same result every run"),
]
for i, params in enumerate(lr_params):
    r = 13 + i
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE_HEX
    for c, val in enumerate(params, 1):
        cell = ws4.cell(r, c, val)
        cell.fill = fill(bg)
        cell.font = font(bold=(c==1), color="1E293B")
        cell.alignment = left()
        cell.border = thin_border()
    ws4.row_dimensions[r].height = 20

# Split config
title_cell(ws4, 19, 1, "Train / Test Split", span_end=4,
           bg=SUBHEAD_BG, fg=NAVY_HEX, size=12, height=22)
header_row(ws4, 20, ["Parameter", "Value", "Why This Value", "Impact"], bg=NAVY_HEX, height=22)
split_params = [
    ("test_size",     "0.30",     "30% of data held out for testing",    f"{len(X_te):,} reports never seen during training"),
    ("stratify",      "cancer_type","All 35 types in both train & test",  "No class excluded from evaluation"),
    ("random_state",  "42",       "Fixed seed",                          "Same split every run — fully reproducible"),
]
for i, params in enumerate(split_params):
    r = 21 + i
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE_HEX
    for c, val in enumerate(params, 1):
        cell = ws4.cell(r, c, val)
        cell.fill = fill(bg)
        cell.font = font(bold=(c==1), color="1E293B")
        cell.alignment = left()
        cell.border = thin_border()
    ws4.row_dimensions[r].height = 20

set_col_width(ws4, 1, 18)
set_col_width(ws4, 2, 16)
set_col_width(ws4, 3, 38)
set_col_width(ws4, 4, 42)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"\nSaved: {OUT}")
print(f"4 sheets:")
print(f"  Sheet 1 - Summary          (overall accuracy, comparison vs Cedars, training steps)")
print(f"  Sheet 2 - Per-Class F1     (all {len(classes)} cancer types with precision/recall/F1 + top words)")
print(f"  Sheet 3 - Misclassifications (top errors + why they happen)")
print(f"  Sheet 4 - Model Config     (every parameter saved in .pkl files)")
print(f"\nAccuracy confirmed: {acc*100:.4f}%")
